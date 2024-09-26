import os, csv, sys, openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter



def init_no_transfer():
    csv_no_transfer = "C:\\Users\\syske\\Downloads\\vod未迁移企业信息.csv"
    f = open(csv_no_transfer, encoding='utf-8' ,errors='ignore')
    reader = csv.reader(f)
    eids = []
    for row_index, row in enumerate(reader):
        print(row[0])
        eids.append(row[0])
    return eids
    
    
def filter_data():
    csv_count = "C:\\Users\\syske\\Downloads\\vod统计信息.csv"
    f = open(csv_count, encoding='utf-8' ,errors='ignore')
    reader = csv.reader(f)
    eids = init_no_transfer()
    out_datas = []
    wb = Workbook()
    # active sheet
    ws = wb.active

    dest_filename = "vod未迁移企业统计信息.xlsx"
    for row_index, row in enumerate(reader):
        if row[0] in eids:
            out_datas.append(row)
    for row_index, row in enumerate(out_datas):
        for column_index, cell in enumerate(row):
            print(column_index)
            column_letter = get_column_letter((column_index + 1))
            ws.cell(row_index + 1, (column_index + 1)).value = cell
            print(f'line:{row_index+1}')
    wb.save(filename = dest_filename)
    print("new Cashflow created")
    
    
    
if __name__ == '__main__':
    filter_data()