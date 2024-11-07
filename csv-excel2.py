import os, csv, sys, openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


#Open an xlsx for reading
wb = Workbook()
# active sheet
ws = wb.active

dest_filename = "vod统计信息.xlsx"
csv_filename = "C:\\Users\\syske\\Downloads\\vod统计信息.csv"

#Copy in csv
f = open(csv_filename, encoding='utf-8' ,errors='ignore')
reader = csv.reader(f)
for row_index, row in enumerate(reader):
    for column_index, cell in enumerate(row):
        column_letter = get_column_letter((column_index + 1))
        ws.cell(row_index + 1, (column_index + 1)).value = cell
    print(f'line:{row_index+1}')

wb.save(filename = dest_filename)

print("new Cashflow created")